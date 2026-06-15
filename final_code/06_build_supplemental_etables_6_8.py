#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Build supplemental eTable 6/7/8 for CO2-rSO2 Model A GAMs.

This script reuses the latest model logic from:
  01_main_gam_analysis.py

Outputs:
  - supplemental_etable6_model_performance_co2_rso2.csv
  - supplemental_etable6_model_performance_co2_rso2.md
  - supplemental_etable7_nonparametric_terms_co2_rso2.csv
  - supplemental_etable8_parametric_terms_co2_rso2.csv
  - Supplemental_eTables6_8_CO2_rSO2.xlsx
"""

import importlib.util
import json
import os
import re
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


LOCAL_MODEL_SCRIPT = Path(__file__).resolve().parent / "01_main_gam_analysis.py"
WORK_DIR = Path(os.getenv(
    "ETABLE68_WORK_DIR",
    "/N/project/waveform_mortality/ZhaoZhang/contour_zhao_all_9_15_2025/GAM_Co2_Scto2_4_19_2026",
))
MODEL_SCRIPT = Path(os.getenv("ETABLE68_MODEL_SCRIPT", str(LOCAL_MODEL_SCRIPT)))

os.environ.setdefault("INTRA5_HEMO_ADJUST_MODE", "map_ci_te")
os.environ.setdefault("INTRA5_SUBSAMPLE_SIZE", "10000")

OUT_ET6 = WORK_DIR / "supplemental_etable6_model_performance_co2_rso2.csv"
OUT_ET6_MD = WORK_DIR / "supplemental_etable6_model_performance_co2_rso2.md"
OUT_ET7 = WORK_DIR / "supplemental_etable7_nonparametric_terms_co2_rso2.csv"
OUT_ET8 = WORK_DIR / "supplemental_etable8_parametric_terms_co2_rso2.csv"
OUT_XLSX = WORK_DIR / "Supplemental_eTables6_8_CO2_rSO2.xlsx"

COHORT_MAP = {
    "rSO2_Ch1": "Left SctO2 cohort",
    "rSO2_Ch2": "Right SctO2 cohort",
    "rSO2_Ch3": "SftO2 cohort",
}
MODEL_LABEL_MAP = {
    "rSO2_Ch1": "Left SctO2 model",
    "rSO2_Ch2": "Right SctO2 model",
    "rSO2_Ch3": "SftO2 model",
}
YCOL_ORDER = ["rSO2_Ch1", "rSO2_Ch2", "rSO2_Ch3"]
ETABLE7_VARIABLE_ORDER = [
    "ET_CO2",
    "TEMP",
    "FiO2_new",
    "MAP x CI",
    "RRtotal",
    "TVinsp",
    "Pmean",
    "Age",
    "BMI",
    "Cardiac_index",
    "Mean_blood_pressure",
    "Hb",
    "Left_SctO2",
    "Right_SctO2",
    "SstO2",
]
ETABLE7_VARIABLE_LABELS = {
    "ET_CO2": "EtCO2",
    "TEMP": "Temperature",
    "FiO2_new": "FiO2",
    "MAP x CI": "MAP-CI interaction",
    "RRtotal": "Respiratory rate",
    "TVinsp": "Tidal volume",
    "Pmean": "PEEP",
    "Cardiac_index": "Preoperative CI",
    "Mean_blood_pressure": "Preoperative MAP",
    "Hb": "Hemoglobin",
    "Left_SctO2": "Preoperative left SctO2",
    "Right_SctO2": "Preoperative right SctO2",
    "SstO2": "Preoperative SftO2",
}
ETABLE7_TERM_LABELS = {
    "Primary smooth": "Penalized spline",
    "Adjustment smooth": "Penalized spline",
    "Tensor product smooth": "Tensor-product penalized spline",
}
ETABLE8_VARIABLE_ORDER = [
    "Smoking_new",
    "Drinking_status",
    "Diabetes_status",
    "Hypertension",
    "Carotid_artery_disease",
    "Statin_1",
    "SEX",
]
ETABLE8_VARIABLE_LABELS = {
    "Smoking_new": "Smoking status",
    "Drinking_status": "Drinking status",
    "Diabetes_status": "Diabetes",
    "Hypertension": "Hypertension",
    "Carotid_artery_disease": "Carotid artery disease",
    "Statin_1": "Statin use",
    "SEX": "Male sex",
}


def load_module(script_path: Path):
    spec = importlib.util.spec_from_file_location("m4_19", script_path)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)  # type: ignore[attr-defined]
    return module


def fmt_p(p: float) -> str:
    if p is None or not np.isfinite(p):
        return "NA"
    if p < 0.001:
        return "<0.001"
    return f"{p:.3f}"


def safe_float(x) -> float:
    try:
        v = float(x)
        return v if np.isfinite(v) else np.nan
    except Exception:
        return np.nan


def sort_etable7(et7: pd.DataFrame) -> pd.DataFrame:
    d = et7.copy()
    model_rank = {k: i for i, k in enumerate(YCOL_ORDER)}
    var_rank = {k: i for i, k in enumerate(ETABLE7_VARIABLE_ORDER)}
    d["model_rank"] = d["ycol"].map(model_rank).fillna(len(model_rank))
    d["variable_rank"] = d["variable"].map(var_rank).fillna(len(var_rank))
    d = d.sort_values(["model_rank", "variable_rank", "variable"], kind="mergesort")
    return d.drop(columns=["model_rank", "variable_rank"])


def etable7_variable_label(variable: str) -> str:
    return ETABLE7_VARIABLE_LABELS.get(str(variable), str(variable))


def etable7_term_label(term_type: str) -> str:
    return ETABLE7_TERM_LABELS.get(str(term_type), str(term_type))


def sort_etable8(et8: pd.DataFrame) -> pd.DataFrame:
    d = et8.copy()
    model_rank = {k: i for i, k in enumerate(YCOL_ORDER)}
    var_rank = {k: i for i, k in enumerate(ETABLE8_VARIABLE_ORDER)}
    d["model_rank"] = d["ycol"].map(model_rank).fillna(len(model_rank))
    d["variable_rank"] = d["variable"].map(var_rank).fillna(len(var_rank))
    d = d.sort_values(["model_rank", "variable_rank", "variable"], kind="mergesort")
    return d.drop(columns=["model_rank", "variable_rank"])


def etable8_variable_label(variable: str) -> str:
    return ETABLE8_VARIABLE_LABELS.get(str(variable), str(variable))


def build_df_base(m) -> pd.DataFrame:
    model_extra = list(getattr(m, "MODEL_EXTRA_LOAD_VARS", getattr(m, "OPTIONAL_INTRAOP_COVARS", [])))
    need_cols = (
        ["stay_id", "patient_ID", "patient_id", "obstime"]
        + list(m.PRIMARY_VARS)
        + model_extra
        + list(m.OUTCOMES)
        + list(m.ADJ_CONT_CAND)
        + list(m.ADJ_CAT_CAND)
        + list(m.SUBGROUP_STATIC_VARS)
        + ["Sex", "SEX", "Age"]
    )
    df_ts = m.read_csv_folder_selected_cached(m.CSV_DIR, m.CSV_GLOB, need_cols=need_cols)
    static_need_cols = sorted(set(list(m.ADJ_CONT_CAND) + list(m.ADJ_CAT_CAND) + list(m.SUBGROUP_STATIC_VARS) + ["Sex", "SEX", "Age"]))
    static_df = m.read_static_selected_multi(
        xlsx_paths=[m.XLSX_PATH_MAIN, m.XLSX_PATH_SUBGROUP],
        need_cols=static_need_cols,
    )

    df_ts = m._coerce_stay_id(df_ts)
    static_df = m._coerce_stay_id(static_df)
    df_ts = m._coerce_patient_id(df_ts)
    static_df = m._coerce_patient_id(static_df)

    merge_key = None
    for k in ["stay_id", "patient_ID", "patient_id"]:
        if (k in df_ts.columns) and (k in static_df.columns):
            merge_key = k
            break
    if merge_key is not None:
        static_cols = [c for c in static_df.columns if c != merge_key]
        df_base = df_ts.merge(static_df[[merge_key] + static_cols], on=merge_key, how="left")
    else:
        df_base = df_ts.copy()

    if "Sex" not in df_base.columns and "SEX" in df_base.columns:
        df_base["Sex"] = df_base["SEX"]
    if "SEX" not in df_base.columns and "Sex" in df_base.columns:
        df_base["SEX"] = df_base["Sex"]

    m.safe_numeric(
        df_base,
        [
            c
            for c in list(m.PRIMARY_VARS)
            + model_extra
            + list(m.OUTCOMES)
            + list(m.ADJ_CONT_CAND)
            + list(m.ADJ_CAT_CAND)
            + list(m.SUBGROUP_STATIC_VARS)
            + ["Sex", "SEX", "Age"]
            if c in df_base.columns
        ],
    )
    m.maybe_convert_fio2_to_percent(df_base, col="FiO2_new")
    return df_base


def get_cached_pool(m, ycol: str, sec: int = 1, subgroup_tag: str = "All", subgroup_query: str = "") -> pd.DataFrame:
    data_source_sig = m.build_data_source_signature()
    cache_key = m.build_pool_cache_key(
        data_source_sig=data_source_sig,
        sec=int(sec),
        subgroup_tag=str(subgroup_tag),
        subgroup_query=str(subgroup_query),
        ycol=str(ycol),
    )
    cache_dir = Path(m.POOL_CACHE_ROOT) / f"intra5_pool_{cache_key}"
    loaded = m._load_pool_cache(cache_dir)
    if loaded is None:
        print(f"[info] pool cache not found for {ycol}, rebuilding: {cache_dir}")
        df_base = build_df_base(m)
        df_sec = m.downsample(df_base, sec=int(sec))
        df_sub = m.apply_subgroup_query(df_sec, subgroup_query) if str(subgroup_query).strip() else df_sec
        pool, _, _, _, _, _, _ = m.build_unified_pool_cached(
            df_sub,
            ycol=str(ycol),
            sec=int(sec),
            subgroup_tag=str(subgroup_tag),
            subgroup_query=str(subgroup_query),
            data_source_sig=data_source_sig,
        )
        return pool
    pool, _, _, _ = loaded
    return pool


def extract_term_feature_index(term) -> Optional[int]:
    if hasattr(term, "feature"):
        feat = getattr(term, "feature")
        if feat is not None:
            try:
                return int(feat)
            except Exception:
                try:
                    arr = np.asarray(feat).reshape(-1)
                    return int(arr[0])
                except Exception:
                    pass
    txt = repr(term)
    m = re.search(r"\((\d+)\)", txt)
    if m:
        return int(m.group(1))
    return None


def make_reference_row(df_model: pd.DataFrame, feature_names: List[str]) -> np.ndarray:
    vals = []
    for c in feature_names:
        s = pd.to_numeric(df_model[c], errors="coerce")
        med = s.median()
        if not np.isfinite(med):
            med = 0.0
        vals.append(float(med))
    return np.asarray(vals, dtype=float).reshape(1, -1)


def factor_single_effect(
    gam,
    cov_mat: np.ndarray,
    x_ref: np.ndarray,
    feat_idx: int,
    observed_codes: np.ndarray,
) -> Dict[str, float]:
    codes = np.asarray(observed_codes, dtype=float)
    codes = codes[np.isfinite(codes)]
    uniq = np.unique(codes)
    if len(uniq) == 0:
        return {"estimate_beta": 0.0, "std_error": 0.0}

    uniq = np.sort(uniq)
    if np.any(np.isclose(uniq, 0.0)) and np.any(np.isclose(uniq, 1.0)):
        ref = 0.0
        alt = 1.0
    elif len(uniq) >= 2:
        ref = float(uniq[0])
        alt = float(uniq[1])
    else:
        ref = float(uniq[0])
        alt = ref

    x0 = x_ref.copy()
    x1 = x_ref.copy()
    x0[0, feat_idx] = ref
    x1[0, feat_idx] = alt

    est = float(gam.predict(x1)[0] - gam.predict(x0)[0])
    lp0 = gam._modelmat(x0).toarray()
    lp1 = gam._modelmat(x1).toarray()
    dx = lp1 - lp0
    var_diff = float(dx @ cov_mat @ dx.T)
    se = float(np.sqrt(max(0.0, var_diff)))
    return {"estimate_beta": est, "std_error": se}


def collect_model_tables(
    gam,
    ycol: str,
    cohort: str,
    model_label: str,
    feature_names: List[str],
    primary_vars: List[str],
    tensor_vars: List[str],
    smooth_covars: List[str],
    cont_cov: List[str],
    cat_cov: List[str],
    df_ref: pd.DataFrame,
    df_eval: Optional[pd.DataFrame],
    model_tag: str,
    n_splines_used,
    lam_used,
) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    st = gam.statistics_ if isinstance(gam.statistics_, dict) else {}
    pvals = np.asarray(st.get("p_values", []), dtype=float)
    se_all = np.asarray(st.get("se", []), dtype=float)
    coef_all = np.asarray(gam.coef_, dtype=float)
    edof_pc = np.asarray(st.get("edof_per_coef", []), dtype=float)
    cov_mat = st.get("cov", None)
    x_ref = make_reference_row(df_ref, feature_names=feature_names)

    X_train = df_ref[feature_names].astype(float).values
    y_train = df_ref[ycol].astype(float).values
    yhat_train = np.asarray(gam.predict(X_train), dtype=float)
    rmse_train = float(np.sqrt(np.mean((y_train - yhat_train) ** 2)))
    mae_train = float(np.mean(np.abs(y_train - yhat_train)))
    ss_res_train = float(np.sum((y_train - yhat_train) ** 2))
    ss_tot_train = float(np.sum((y_train - np.mean(y_train)) ** 2))
    r2_train = 1.0 - ss_res_train / ss_tot_train if ss_tot_train > 0 else np.nan

    df_eval_use = df_eval if df_eval is not None else df_ref
    X_eval = df_eval_use[feature_names].astype(float).values
    y_eval = df_eval_use[ycol].astype(float).values
    yhat_eval = np.asarray(gam.predict(X_eval), dtype=float)
    rmse_eval = float(np.sqrt(np.mean((y_eval - yhat_eval) ** 2)))
    mae_eval = float(np.mean(np.abs(y_eval - yhat_eval)))
    ss_res_eval = float(np.sum((y_eval - yhat_eval) ** 2))
    ss_tot_eval = float(np.sum((y_eval - np.mean(y_eval)) ** 2))
    r2_eval = 1.0 - ss_res_eval / ss_tot_eval if ss_tot_eval > 0 else np.nan

    pr2 = st.get("pseudo_r2", {})
    if isinstance(pr2, dict):
        dev_exp = safe_float(pr2.get("explained_deviance"))
        mc_r2 = safe_float(pr2.get("McFadden"))
    else:
        dev_exp = safe_float(pr2)
        mc_r2 = np.nan

    fit_rows = [
        ("Samples, n", safe_float(st.get("n_samples"))),
        ("Features, n", safe_float(st.get("m_features"))),
        ("Effective degrees of freedom (model)", safe_float(st.get("edof"))),
        ("Deviance explained", dev_exp),
        ("McFadden R2", mc_r2),
        ("AIC", safe_float(st.get("AIC"))),
        ("AICc", safe_float(st.get("AICc"))),
        ("GCV", safe_float(st.get("GCV"))),
        ("UBRE", safe_float(st.get("UBRE"))),
        ("Log Likelihood", safe_float(st.get("loglikelihood"))),
        ("Scale", safe_float(st.get("scale"))),
        ("Training rows, n", float(len(df_ref))),
        ("RMSE (training sample)", rmse_train),
        ("MAE (training sample)", mae_train),
        ("Traditional R2 (training sample)", r2_train),
        ("Evaluation rows, n (full analytic pool)", float(len(df_eval_use))),
        ("RMSE (full analytic pool)", rmse_eval),
        ("MAE (full analytic pool)", mae_eval),
        ("Traditional R2 (full analytic pool)", r2_eval),
    ]
    et6 = pd.DataFrame(
        [
            {
                "ycol": ycol,
                "cohort": cohort,
                "model": model_label,
                "model_tag": model_tag,
                "n_splines_main_used_json": json.dumps(n_splines_used, ensure_ascii=False),
                "lam_used_json": json.dumps(lam_used, ensure_ascii=False),
                "metric": k,
                "value": v,
            }
            for k, v in fit_rows
        ]
    )

    rows7 = []
    rows8 = []
    primary_idx = set(range(len(primary_vars)))
    smooth_cov_idx = set(
        range(
            len(primary_vars) + len(tensor_vars),
            len(primary_vars) + len(tensor_vars) + len(smooth_covars),
        )
    )
    coef_idx = 0
    for ti, term in enumerate(gam.terms):
        nc = int(getattr(term, "n_coefs", 0))
        term_type = term.__class__.__name__
        pval = float(pvals[ti]) if ti < len(pvals) else np.nan
        feat_idx = extract_term_feature_index(term)
        var_name = feature_names[feat_idx] if (feat_idx is not None and feat_idx < len(feature_names)) else "Intercept"
        c_slice = coef_all[coef_idx : coef_idx + nc] if coef_idx + nc <= len(coef_all) else np.array([])
        se_slice = se_all[coef_idx : coef_idx + nc] if coef_idx + nc <= len(se_all) else np.array([])
        ed_slice = edof_pc[coef_idx : coef_idx + nc] if coef_idx + nc <= len(edof_pc) else np.array([])
        edof_term = float(np.sum(ed_slice)) if len(ed_slice) else np.nan

        if term_type in ("SplineTerm", "TensorTerm"):
            if term_type == "TensorTerm":
                pair_start = len(primary_vars)
                pair_i = max(0, feat_idx - pair_start) if feat_idx is not None else 0
                pair_i = int(pair_i // 2)
                pair_vars = tensor_vars[(2 * pair_i):(2 * pair_i + 2)]
                var_name = " x ".join(pair_vars) if len(pair_vars) == 2 else var_name
                display_type = "Tensor product smooth"
            elif feat_idx in primary_idx:
                display_type = "Primary smooth"
            elif feat_idx in smooth_cov_idx:
                display_type = "Adjustment smooth"
            else:
                display_type = "Smooth spline"
            chi2 = np.nan
            fstat = np.nan
            if cov_mat is not None and len(c_slice) == nc and nc > 0:
                try:
                    block = cov_mat[coef_idx : coef_idx + nc, coef_idx : coef_idx + nc]
                    block_inv = np.linalg.pinv(block)
                    chi2 = float(c_slice @ block_inv @ c_slice)
                    if np.isfinite(edof_term) and edof_term > 0:
                        fstat = float(chi2 / edof_term)
                except Exception:
                    pass
            rows7.append(
                {
                    "ycol": ycol,
                    "cohort": cohort,
                    "model": model_label,
                    "model_tag": model_tag,
                    "variable": var_name,
                    "term_type": display_type,
                    "n_basis_functions": nc,
                    "effect_degrees_of_freedom": edof_term,
                    "chi2_statistic": chi2,
                    "f_statistic": fstat,
                    "p_value": pval,
                    "p_value_display": fmt_p(pval),
                }
            )
        elif term_type == "LinearTerm":
            est = float(c_slice[0]) if len(c_slice) else np.nan
            se = float(se_slice[0]) if len(se_slice) else np.nan
            chi2 = np.nan
            fstat = np.nan
            if np.isfinite(est) and np.isfinite(se) and se > 0:
                chi2 = float((est / se) ** 2)
                if np.isfinite(edof_term) and edof_term > 0:
                    fstat = float(chi2 / edof_term)
            rows7.append(
                {
                    "ycol": ycol,
                    "cohort": cohort,
                    "model": model_label,
                    "model_tag": model_tag,
                    "variable": var_name,
                    "term_type": "Linear",
                    "n_basis_functions": 1,
                    "effect_degrees_of_freedom": edof_term,
                    "chi2_statistic": chi2,
                    "f_statistic": fstat,
                    "p_value": pval,
                    "p_value_display": fmt_p(pval),
                }
            )
        elif term_type == "FactorTerm":
            est = np.nan
            se = np.nan
            if cov_mat is not None and feat_idx is not None and feat_idx < len(feature_names):
                try:
                    eff = factor_single_effect(
                        gam=gam,
                        cov_mat=np.asarray(cov_mat, dtype=float),
                        x_ref=x_ref,
                        feat_idx=int(feat_idx),
                        observed_codes=pd.to_numeric(df_ref[feature_names[feat_idx]], errors="coerce").values,
                    )
                    est = float(eff["estimate_beta"])
                    se = float(eff["std_error"])
                except Exception:
                    est = float(c_slice[0]) if len(c_slice) else np.nan
                    se = float(se_slice[0]) if len(se_slice) else np.nan
            else:
                est = float(c_slice[0]) if len(c_slice) else np.nan
                se = float(se_slice[0]) if len(se_slice) else np.nan
            rows8.append(
                {
                    "ycol": ycol,
                    "cohort": cohort,
                    "model": model_label,
                    "model_tag": model_tag,
                    "variable": var_name,
                    "term_type": "Factor",
                    "effect_degrees_of_freedom": edof_term,
                    "estimate_beta": est,
                    "std_error": se,
                    "p_value": pval,
                    "p_value_display": fmt_p(pval),
                }
            )

        coef_idx += nc

    et7 = pd.DataFrame(rows7)
    et8 = pd.DataFrame(rows8)
    return et6, et7, et8


def write_xlsx(et6: pd.DataFrame, et7: pd.DataFrame, et8: pd.DataFrame, out_xlsx: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    title_font = Font(name="Calibri", size=12, bold=True)
    header_font = Font(name="Calibri", size=10, bold=True)
    body_font = Font(name="Calibri", size=10)
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill = PatternFill(fill_type="solid", fgColor="F2F2F2")
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # eTable 6 (wide)
    ws6 = wb.create_sheet("eTable6_model_fit")
    ws6.merge_cells("A1:D1")
    ws6["A1"] = "eTable 6. Model performance and goodness-of-fit statistics for Model A generalized additive models"
    ws6["A1"].font = title_font
    ws6["A1"].alignment = left
    header6 = ["Metric", "Left SctO2 cohort", "Right SctO2 cohort", "SftO2 cohort"]
    for c, v in enumerate(header6, start=1):
        cell = ws6.cell(3, c, v)
        cell.font = header_font
        cell.fill = fill
        cell.border = border
        cell.alignment = left if c == 1 else center

    metric_order = [
        "Samples, n",
        "Features, n",
        "Effective degrees of freedom (model)",
        "Deviance explained",
        "McFadden R2",
        "AIC",
        "AICc",
        "GCV",
        "UBRE",
        "Log Likelihood",
        "Scale",
        "Training rows, n",
        "RMSE (training sample)",
        "MAE (training sample)",
        "Traditional R2 (training sample)",
        "Evaluation rows, n (full analytic pool)",
        "RMSE (full analytic pool)",
        "MAE (full analytic pool)",
        "Traditional R2 (full analytic pool)",
    ]
    p6 = et6.pivot_table(index="metric", columns="cohort", values="value", aggfunc="first")
    r = 4
    for metric in metric_order:
        ws6.cell(r, 1, metric).font = body_font
        ws6.cell(r, 1).alignment = left
        ws6.cell(r, 1).border = border
        for j, cohort in enumerate(["Left SctO2 cohort", "Right SctO2 cohort", "SftO2 cohort"], start=2):
            val = p6.loc[metric, cohort] if metric in p6.index and cohort in p6.columns else np.nan
            txt = ""
            if np.isfinite(safe_float(val)):
                txt = f"{float(val):,.6f}" if abs(float(val)) < 1000 else f"{float(val):,.2f}"
            ws6.cell(r, j, txt).font = body_font
            ws6.cell(r, j).alignment = center
            ws6.cell(r, j).border = border
        r += 1
    ws6.column_dimensions["A"].width = 38
    ws6.column_dimensions["B"].width = 24
    ws6.column_dimensions["C"].width = 24
    ws6.column_dimensions["D"].width = 24

    # eTable 7
    ws7 = wb.create_sheet("eTable7_continuous")
    ws7.merge_cells("A1:G1")
    ws7["A1"] = "eTable 7. Continuous term estimates for Model A generalized additive models"
    ws7["A1"].font = title_font
    ws7["A1"].alignment = left
    header7 = ["Model", "Variable", "Term Type", "N basis functions", "Effect degrees of freedom", "F statistic", "P value"]
    for c, v in enumerate(header7, start=1):
        cell = ws7.cell(3, c, v)
        cell.font = header_font
        cell.fill = fill
        cell.border = border
        cell.alignment = center if c >= 4 else left
    r = 4
    et7s = sort_etable7(et7)
    for _, row in et7s.iterrows():
        vals = [
            row["model"],
            etable7_variable_label(row["variable"]),
            etable7_term_label(row["term_type"]),
            int(row["n_basis_functions"]) if np.isfinite(safe_float(row["n_basis_functions"])) else "",
            f"{float(row['effect_degrees_of_freedom']):.2f}" if np.isfinite(safe_float(row["effect_degrees_of_freedom"])) else "",
            f"{float(row['f_statistic']):.2f}" if np.isfinite(safe_float(row["f_statistic"])) else "",
            fmt_p(safe_float(row["p_value"])),
        ]
        for c, v in enumerate(vals, start=1):
            cell = ws7.cell(r, c, v)
            cell.font = body_font
            cell.border = border
            cell.alignment = center if c >= 4 else left
        r += 1
    for col, width in {"A": 24, "B": 28, "C": 32, "D": 18, "E": 24, "F": 14, "G": 10}.items():
        ws7.column_dimensions[col].width = width

    # eTable 8
    ws8 = wb.create_sheet("eTable8_categorical")
    ws8.merge_cells("A1:G1")
    ws8["A1"] = "eTable 8. Categorical term estimates for Model A generalized additive models"
    ws8["A1"].font = title_font
    ws8["A1"].alignment = left
    header8 = ["Model", "Variable", "Term Type", "Effect degrees of freedom", "Estimate (β)", "Std error", "P value"]
    for c, v in enumerate(header8, start=1):
        cell = ws8.cell(3, c, v)
        cell.font = header_font
        cell.fill = fill
        cell.border = border
        cell.alignment = center if c >= 4 else left
    r = 4
    et8s = sort_etable8(et8)
    for _, row in et8s.iterrows():
        vals = [
            row["model"],
            etable8_variable_label(row["variable"]),
            row["term_type"],
            f"{float(row['effect_degrees_of_freedom']):.2f}" if np.isfinite(safe_float(row["effect_degrees_of_freedom"])) else "",
            f"{float(row['estimate_beta']):.3f}" if np.isfinite(safe_float(row["estimate_beta"])) else "",
            f"{float(row['std_error']):.3f}" if np.isfinite(safe_float(row["std_error"])) else "",
            fmt_p(safe_float(row["p_value"])),
        ]
        for c, v in enumerate(vals, start=1):
            cell = ws8.cell(r, c, v)
            cell.font = body_font
            cell.border = border
            cell.alignment = center if c >= 4 else left
        r += 1
    for col, width in {"A": 24, "B": 24, "C": 16, "D": 24, "E": 14, "F": 14, "G": 10}.items():
        ws8.column_dimensions[col].width = width

    wb.save(out_xlsx)


def _metric_value(p6: pd.DataFrame, metric: str, cohort: str) -> float:
    if metric not in p6.index or cohort not in p6.columns:
        return np.nan
    return safe_float(p6.loc[metric, cohort])


def _fmt_metric_value(value: float, kind: str) -> str:
    if not np.isfinite(safe_float(value)):
        return "-"
    value = float(value)
    if kind == "int":
        return f"{value:,.0f}"
    if kind == "edf":
        return f"{value:.2f}"
    if kind == "aic":
        return f"{value:,.0f}"
    if kind == "pct":
        return f"{value * 100:.1f}"
    if kind == "r2":
        return f"{value:.3f}"
    if kind == "err":
        return f"{value:.2f}"
    return f"{value:.3f}"


def write_etable6_markdown(et6: pd.DataFrame, out_md: Path) -> None:
    p6 = et6.pivot_table(index="metric", columns="cohort", values="value", aggfunc="first")
    cohorts = ["Left SctO2 cohort", "Right SctO2 cohort", "SftO2 cohort"]
    header = ["Metric", "Left SctO2 Model", "Right SctO2 Model", "SftO2 Model"]
    rows = [
        ("N Samples", "Samples, n", "int"),
        ("N Features", "Features, n", "int"),
        ("Effect Degrees of Freedom", "Effective degrees of freedom (model)", "edf"),
        ("AIC", "AIC", "aic"),
        ("GCV", "GCV", "num"),
        ("Scale Parameter", "Scale", "num"),
        ("Deviance Explained (%)", "Deviance explained", "pct"),
        ("R2", "Traditional R2 (training sample)", "r2"),
        ("RMSE", "RMSE (training sample)", "err"),
        ("MAE", "MAE (training sample)", "err"),
    ]

    lines = [
        "## eTable 6. Model performance and goodness-of-fit statistics for Model A generalized additive models of left cerebral, right cerebral, and forearm tissue oxygenation",
        "",
        "| " + " | ".join(header) + " |",
        "| --- | --- | --- | --- |",
    ]
    for display, source, kind in rows:
        vals = [_fmt_metric_value(_metric_value(p6, source, cohort), kind) for cohort in cohorts]
        lines.append("| " + " | ".join([display] + vals) + " |")
    lines.extend(
        [
            "",
            "Notes:",
            "- Metrics are calculated from the 10,000-row Model A fitting sample unless otherwise specified.",
            "- Model A included smooth terms for EtCO2, temperature, and FiO2, tensor-product MAP-CI adjustment, smooth respiratory covariates, and baseline covariate adjustment.",
            "- SctO2, cerebral tissue oxygen saturation; SftO2, forearm tissue oxygen saturation; EtCO2, end-tidal carbon dioxide; FiO2, fraction of inspired oxygen; MAP, mean arterial pressure; CI, cardiac index; AIC, Akaike information criterion; GCV, generalized cross-validation; RMSE, root mean square error; MAE, mean absolute error.",
            "",
        ]
    )
    out_md.write_text("\n".join(lines), encoding="utf-8")


def main():
    if not MODEL_SCRIPT.exists():
        raise FileNotFoundError(f"Model script not found: {MODEL_SCRIPT}")

    m = load_module(MODEL_SCRIPT)
    outcomes_env = [x.strip() for x in os.getenv("ETABLE68_OUTCOMES", "").split(",") if x.strip()]
    ycols = [y for y in outcomes_env if y in YCOL_ORDER] if outcomes_env else list(YCOL_ORDER)
    if len(ycols) < len(YCOL_ORDER):
        print(f"[warn] ETABLE68_OUTCOMES override active, only running cohorts: {ycols}")
    sample_n = int(os.getenv("ETABLE68_SUBSAMPLE_N", "10000"))
    auto_tune = os.getenv("ETABLE68_AUTO_TUNE", "1" if bool(m.AUTO_TUNE_SMOOTH) else "0") == "1"
    replace = os.getenv("ETABLE68_REPLACE", "1" if bool(m.REF_SAMPLE_REPLACE) else "0") == "1"

    et6_all: List[pd.DataFrame] = []
    et7_all: List[pd.DataFrame] = []
    et8_all: List[pd.DataFrame] = []

    print(
        f"[cfg] model=Model A te(MAP,CI), sample_n={sample_n}, auto_tune={int(auto_tune)}, "
        f"replace={int(replace)}, outcomes={ycols}"
    )
    for ycol in ycols:
        print(f"[fit] {ycol}")
        pool = get_cached_pool(m, ycol=ycol, sec=1, subgroup_tag="All", subgroup_query="")
        model_tag = "modelA_map_ci_te"

        d_model, pvars, tensor_vars, smooth_covars, cont_cov, cat_cov = m.prepare_model_df(
            pool,
            ycol=ycol,
            extra_exclude_covars=[],
        )
        if len(d_model) < 500:
            raise RuntimeError(f"Insufficient rows for {ycol}: {len(d_model)}")
        if list(tensor_vars) != ["MAP", "CI"]:
            raise RuntimeError(f"Model A requires tensor_vars=['MAP', 'CI']; got {tensor_vars} for {ycol}")

        take_n = min(int(sample_n), len(d_model))
        df_ref = d_model.sample(n=take_n, replace=replace, random_state=int(m.SEED + 17))
        if len(df_ref) < 500:
            raise RuntimeError(f"Reference sample too small for {ycol}: {len(df_ref)}")

        if auto_tune:
            tuned = m.tune_smoothing(
                df_ref,
                ycol=ycol,
                primary_vars=pvars,
                tensor_vars=tensor_vars,
                smooth_covars=smooth_covars,
                cont_cov=cont_cov,
                cat_cov=cat_cov,
            )
            gam = tuned["gam"]
            n_splines_used = tuned.get("n_splines_main", [])
            lam_used = tuned.get("lam", None)
        else:
            gam = m.fit_gam(
                df_model=df_ref,
                ycol=ycol,
                primary_vars=pvars,
                tensor_vars=tensor_vars,
                smooth_covars=smooth_covars,
                cont_cov=cont_cov,
                cat_cov=cat_cov,
                n_splines_main=int(m.N_SPLINES_MAIN),
                lam_fixed=float(m.LAM_FIXED),
            )
            n_splines_used = (
                [int(m.N_SPLINES_MAIN)] * len(pvars)
                + [int(m.N_SPLINES_TENSOR)] * int(len(tensor_vars) / 2)
                + [int(m.N_SPLINES_COV)] * len(smooth_covars)
            )
            lam_used = float(m.LAM_FIXED)

        feature_names = m.build_feature_names(pvars, tensor_vars, smooth_covars, cont_cov, cat_cov)
        cohort = COHORT_MAP.get(ycol, ycol)
        model_label = MODEL_LABEL_MAP.get(ycol, ycol)
        et6, et7, et8 = collect_model_tables(
            gam=gam,
            ycol=ycol,
            cohort=cohort,
            model_label=model_label,
            feature_names=feature_names,
            primary_vars=pvars,
            tensor_vars=tensor_vars,
            smooth_covars=smooth_covars,
            cont_cov=cont_cov,
            cat_cov=cat_cov,
            df_ref=df_ref,
            df_eval=d_model,
            model_tag=model_tag,
            n_splines_used=n_splines_used,
            lam_used=lam_used,
        )
        et6_all.append(et6)
        et7_all.append(et7)
        et8_all.append(et8)
        print(f"[done] {ycol} rows: et6={len(et6)} et7={len(et7)} et8={len(et8)}")

    d6 = pd.concat(et6_all, ignore_index=True) if et6_all else pd.DataFrame()
    d7 = pd.concat(et7_all, ignore_index=True) if et7_all else pd.DataFrame()
    if not d7.empty:
        d7 = sort_etable7(d7)
    d8 = pd.concat(et8_all, ignore_index=True) if et8_all else pd.DataFrame()
    if not d8.empty:
        d8 = sort_etable8(d8)

    d6.to_csv(OUT_ET6, index=False)
    d7.to_csv(OUT_ET7, index=False)
    d8.to_csv(OUT_ET8, index=False)
    write_xlsx(d6, d7, d8, OUT_XLSX)
    write_etable6_markdown(d6, OUT_ET6_MD)

    print("[write]")
    print(f" - {OUT_ET6}")
    print(f" - {OUT_ET6_MD}")
    print(f" - {OUT_ET7}")
    print(f" - {OUT_ET8}")
    print(f" - {OUT_XLSX}")


if __name__ == "__main__":
    main()
