#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Build supplemental eTable 6/7/8 for CO2-rSO2 GAM models.

This script reuses the latest model logic from:
  contour_4_19_2026_intraop10_totalonly_noetco2_subgroup.py

Outputs:
  - supplemental_etable6_model_performance_co2_rso2.csv
  - supplemental_etable7_nonparametric_terms_co2_rso2.csv
  - supplemental_etable8_parametric_terms_co2_rso2.csv
  - Supplemental_eTables6_8_CO2_rSO2.xlsx
"""

import importlib.util
import json
import math
import os
import re
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


WORK_DIR = Path("/N/project/waveform_mortality/ZhaoZhang/contour_zhao_all_9_15_2025/GAM_Co2_Scto2_4_19_2026")
MODEL_SCRIPT = WORK_DIR / "contour_4_19_2026_intraop10_totalonly_noetco2_subgroup.py"

OUT_ET6 = WORK_DIR / "supplemental_etable6_model_performance_co2_rso2.csv"
OUT_ET7 = WORK_DIR / "supplemental_etable7_nonparametric_terms_co2_rso2.csv"
OUT_ET8 = WORK_DIR / "supplemental_etable8_parametric_terms_co2_rso2.csv"
OUT_XLSX = WORK_DIR / "Supplemental_eTables6_8_CO2_rSO2.xlsx"

COHORT_MAP = {
    "rSO2_Ch1": "Left SctO2 cohort",
    "rSO2_Ch2": "Right SctO2 cohort",
    "rSO2_Ch3": "SftO2 cohort",
}
MODEL_LABEL_MAP = {
    "rSO2_Ch1": "Left SctO2 model",
    "rSO2_Ch2": "Right SctO2 model",
    "rSO2_Ch3": "SftO2 model",
}
YCOL_ORDER = ["rSO2_Ch1", "rSO2_Ch2", "rSO2_Ch3"]


def load_module(script_path: Path):
    spec = importlib.util.spec_from_file_location("m4_19", script_path)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)  # type: ignore[attr-defined]
    return module


def fmt_p(p: float) -> str:
    if p is None or not np.isfinite(p):
        return "NA"
    if p < 0.001:
        return "<0.001"
    return f"{p:.3f}"


def wald_p_from_estimate(est: float, se: float) -> float:
    if not (np.isfinite(est) and np.isfinite(se)) or se <= 0:
        return np.nan
    z = abs(est / se)
    return 2.0 * (1.0 - 0.5 * (1.0 + math.erf(z / math.sqrt(2.0))))


def safe_float(x) -> float:
    try:
        v = float(x)
        return v if np.isfinite(v) else np.nan
    except Exception:
        return np.nan


def choose_model_spec(m, pool: pd.DataFrame) -> Dict[str, object]:
    specs = m.build_model_specs(list(pool.columns), lag_cols_available=[c for c in pool.columns if "_lag" in str(c)])
    if not specs:
        raise RuntimeError("No valid model spec found for pooled data.")
    for sp in specs:
        if str(sp.get("effect_type", "")) == "total":
            return sp
    return specs[0]


def build_df_base(m) -> pd.DataFrame:
    need_cols = (
        ["stay_id", "patient_ID", "patient_id", "obstime"]
        + list(m.PRIMARY_VARS)
        + list(m.MODEL_EXTRA_LOAD_VARS)
        + list(m.OUTCOMES)
        + list(m.ADJ_CONT_CAND)
        + list(m.ADJ_CAT_CAND)
        + list(m.SUBGROUP_STATIC_VARS)
        + ["Sex", "SEX", "Age"]
    )
    df_ts = m.read_csv_folder_selected_cached(m.CSV_DIR, m.CSV_GLOB, need_cols=need_cols)
    static_need_cols = sorted(set(list(m.ADJ_CONT_CAND) + list(m.ADJ_CAT_CAND) + list(m.SUBGROUP_STATIC_VARS) + ["Sex", "SEX", "Age"]))
    static_df = m.read_static_selected_multi(
        xlsx_paths=[m.XLSX_PATH_MAIN, m.XLSX_PATH_SUBGROUP],
        need_cols=static_need_cols,
    )

    df_ts = m._coerce_stay_id(df_ts)
    static_df = m._coerce_stay_id(static_df)
    df_ts = m._coerce_patient_id(df_ts)
    static_df = m._coerce_patient_id(static_df)

    merge_key = None
    for k in ["stay_id", "patient_ID", "patient_id"]:
        if (k in df_ts.columns) and (k in static_df.columns):
            merge_key = k
            break
    if merge_key is not None:
        static_cols = [c for c in static_df.columns if c != merge_key]
        df_base = df_ts.merge(static_df[[merge_key] + static_cols], on=merge_key, how="left")
    else:
        df_base = df_ts.copy()

    if "Sex" not in df_base.columns and "SEX" in df_base.columns:
        df_base["Sex"] = df_base["SEX"]
    if "SEX" not in df_base.columns and "Sex" in df_base.columns:
        df_base["SEX"] = df_base["Sex"]

    m.safe_numeric(
        df_base,
        [
            c
            for c in list(m.PRIMARY_VARS)
            + list(m.MODEL_EXTRA_LOAD_VARS)
            + list(m.OUTCOMES)
            + list(m.ADJ_CONT_CAND)
            + list(m.ADJ_CAT_CAND)
            + list(m.SUBGROUP_STATIC_VARS)
            + ["Sex", "SEX", "Age"]
            if c in df_base.columns
        ],
    )
    m.maybe_convert_fio2_to_percent(df_base, col="FiO2_new")
    return df_base


def get_cached_pool(m, ycol: str, sec: int = 1, subgroup_tag: str = "All", subgroup_query: str = "") -> pd.DataFrame:
    data_source_sig = m.build_data_source_signature()
    cache_key = m.build_pool_cache_key(
        data_source_sig=data_source_sig,
        sec=int(sec),
        subgroup_tag=str(subgroup_tag),
        subgroup_query=str(subgroup_query),
        ycol=str(ycol),
    )
    cache_dir = Path(m.POOL_CACHE_ROOT) / f"intra5_pool_{cache_key}"
    loaded = m._load_pool_cache(cache_dir)
    if loaded is None:
        print(f"[info] pool cache not found for {ycol}, rebuilding: {cache_dir}")
        df_base = build_df_base(m)
        df_sec = m.downsample(df_base, sec=int(sec))
        df_sub = m.apply_subgroup_query(df_sec, subgroup_query) if str(subgroup_query).strip() else df_sec
        pool, _, _, _, _, _, _ = m.build_unified_pool_cached(
            df_sub,
            ycol=str(ycol),
            sec=int(sec),
            subgroup_tag=str(subgroup_tag),
            subgroup_query=str(subgroup_query),
            data_source_sig=data_source_sig,
        )
        return pool
    pool, _, _, _ = loaded
    return pool


def extract_term_feature_index(term) -> Optional[int]:
    if hasattr(term, "feature"):
        feat = getattr(term, "feature")
        if feat is not None:
            try:
                return int(feat)
            except Exception:
                pass
    txt = repr(term)
    m = re.search(r"\((\d+)\)", txt)
    if m:
        return int(m.group(1))
    return None


def make_reference_row(df_model: pd.DataFrame, feature_names: List[str]) -> np.ndarray:
    vals = []
    for c in feature_names:
        s = pd.to_numeric(df_model[c], errors="coerce")
        med = s.median()
        if not np.isfinite(med):
            med = 0.0
        vals.append(float(med))
    return np.asarray(vals, dtype=float).reshape(1, -1)


def factor_single_effect(
    gam,
    cov_mat: np.ndarray,
    x_ref: np.ndarray,
    feat_idx: int,
    observed_codes: np.ndarray,
) -> Dict[str, float]:
    codes = np.asarray(observed_codes, dtype=float)
    codes = codes[np.isfinite(codes)]
    uniq = np.unique(codes)
    if len(uniq) == 0:
        return {"estimate_beta": 0.0, "std_error": 0.0}

    uniq = np.sort(uniq)
    if np.any(np.isclose(uniq, 0.0)) and np.any(np.isclose(uniq, 1.0)):
        ref = 0.0
        alt = 1.0
    elif len(uniq) >= 2:
        ref = float(uniq[0])
        alt = float(uniq[1])
    else:
        ref = float(uniq[0])
        alt = ref

    x0 = x_ref.copy()
    x1 = x_ref.copy()
    x0[0, feat_idx] = ref
    x1[0, feat_idx] = alt

    est = float(gam.predict(x1)[0] - gam.predict(x0)[0])
    lp0 = gam._modelmat(x0).toarray()
    lp1 = gam._modelmat(x1).toarray()
    dx = lp1 - lp0
    var_diff = float(dx @ cov_mat @ dx.T)
    se = float(np.sqrt(max(0.0, var_diff)))
    return {"estimate_beta": est, "std_error": se}


def collect_model_tables(
    gam,
    ycol: str,
    cohort: str,
    model_label: str,
    feature_names: List[str],
    primary_vars: List[str],
    cont_cov: List[str],
    cat_cov: List[str],
    df_ref: pd.DataFrame,
    df_eval: Optional[pd.DataFrame],
    model_tag: str,
    n_splines_used,
    lam_used,
) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    st = gam.statistics_ if isinstance(gam.statistics_, dict) else {}
    pvals = np.asarray(st.get("p_values", []), dtype=float)
    se_all = np.asarray(st.get("se", []), dtype=float)
    coef_all = np.asarray(gam.coef_, dtype=float)
    edof_pc = np.asarray(st.get("edof_per_coef", []), dtype=float)
    cov_mat = st.get("cov", None)
    x_ref = make_reference_row(df_ref, feature_names=feature_names)

    X_train = df_ref[feature_names].astype(float).values
    y_train = df_ref[ycol].astype(float).values
    yhat_train = np.asarray(gam.predict(X_train), dtype=float)
    rmse_train = float(np.sqrt(np.mean((y_train - yhat_train) ** 2)))
    ss_res_train = float(np.sum((y_train - yhat_train) ** 2))
    ss_tot_train = float(np.sum((y_train - np.mean(y_train)) ** 2))
    r2_train = 1.0 - ss_res_train / ss_tot_train if ss_tot_train > 0 else np.nan

    df_eval_use = df_eval if df_eval is not None else df_ref
    X_eval = df_eval_use[feature_names].astype(float).values
    y_eval = df_eval_use[ycol].astype(float).values
    yhat_eval = np.asarray(gam.predict(X_eval), dtype=float)
    rmse_eval = float(np.sqrt(np.mean((y_eval - yhat_eval) ** 2)))
    ss_res_eval = float(np.sum((y_eval - yhat_eval) ** 2))
    ss_tot_eval = float(np.sum((y_eval - np.mean(y_eval)) ** 2))
    r2_eval = 1.0 - ss_res_eval / ss_tot_eval if ss_tot_eval > 0 else np.nan

    pr2 = st.get("pseudo_r2", {})
    if isinstance(pr2, dict):
        dev_exp = safe_float(pr2.get("explained_deviance"))
        mc_r2 = safe_float(pr2.get("McFadden"))
    else:
        dev_exp = safe_float(pr2)
        mc_r2 = np.nan

    fit_rows = [
        ("N Samples", safe_float(st.get("n_samples"))),
        ("N Features", safe_float(st.get("m_features"))),
        ("Effective DOF (model)", safe_float(st.get("edof"))),
        ("Deviance Explained", dev_exp),
        ("McFadden R2", mc_r2),
        ("AIC", safe_float(st.get("AIC"))),
        ("AICc", safe_float(st.get("AICc"))),
        ("GCV", safe_float(st.get("GCV"))),
        ("UBRE", safe_float(st.get("UBRE"))),
        ("Log Likelihood", safe_float(st.get("loglikelihood"))),
        ("Scale", safe_float(st.get("scale"))),
        ("N Training Rows", float(len(df_ref))),
        ("RMSE (training sample)", rmse_train),
        ("R2 Traditional (training sample)", r2_train),
        ("N Eval Rows (full analytic pool)", float(len(df_eval_use))),
        ("RMSE (full analytic pool)", rmse_eval),
        ("R2 Traditional (full analytic pool)", r2_eval),
    ]
    et6 = pd.DataFrame(
        [
            {
                "ycol": ycol,
                "cohort": cohort,
                "model": model_label,
                "model_tag": model_tag,
                "n_splines_main_used_json": json.dumps(n_splines_used, ensure_ascii=False),
                "lam_used_json": json.dumps(lam_used, ensure_ascii=False),
                "metric": k,
                "value": v,
            }
            for k, v in fit_rows
        ]
    )

    rows7 = []
    rows8 = []
    coef_idx = 0
    for ti, term in enumerate(gam.terms):
        nc = int(getattr(term, "n_coefs", 0))
        term_type = term.__class__.__name__
        pval = float(pvals[ti]) if ti < len(pvals) else np.nan
        feat_idx = extract_term_feature_index(term)
        var_name = feature_names[feat_idx] if (feat_idx is not None and feat_idx < len(feature_names)) else "Intercept"
        c_slice = coef_all[coef_idx : coef_idx + nc] if coef_idx + nc <= len(coef_all) else np.array([])
        se_slice = se_all[coef_idx : coef_idx + nc] if coef_idx + nc <= len(se_all) else np.array([])
        ed_slice = edof_pc[coef_idx : coef_idx + nc] if coef_idx + nc <= len(edof_pc) else np.array([])
        edof_term = float(np.sum(ed_slice)) if len(ed_slice) else np.nan

        if term_type in ("SplineTerm", "TensorTerm"):
            chi2 = np.nan
            fstat = np.nan
            if cov_mat is not None and len(c_slice) == nc and nc > 0:
                try:
                    block = cov_mat[coef_idx : coef_idx + nc, coef_idx : coef_idx + nc]
                    block_inv = np.linalg.pinv(block)
                    chi2 = float(c_slice @ block_inv @ c_slice)
                    if np.isfinite(edof_term) and edof_term > 0:
                        fstat = float(chi2 / edof_term)
                except Exception:
                    pass
            rows7.append(
                {
                    "ycol": ycol,
                    "cohort": cohort,
                    "model": model_label,
                    "model_tag": model_tag,
                    "variable": var_name,
                    "term_type": "Smooth Spline" if term_type == "SplineTerm" else "Tensor Product",
                    "n_basis_functions": nc,
                    "effect_degrees_of_freedom": edof_term,
                    "chi2_statistic": chi2,
                    "f_statistic": fstat,
                    "p_value": pval,
                    "p_value_display": fmt_p(pval),
                }
            )
        elif term_type == "LinearTerm":
            est = float(c_slice[0]) if len(c_slice) else np.nan
            se = float(se_slice[0]) if len(se_slice) else np.nan
            chi2 = np.nan
            fstat = np.nan
            if np.isfinite(est) and np.isfinite(se) and se > 0:
                chi2 = float((est / se) ** 2)
                if np.isfinite(edof_term) and edof_term > 0:
                    fstat = float(chi2 / edof_term)
            rows7.append(
                {
                    "ycol": ycol,
                    "cohort": cohort,
                    "model": model_label,
                    "model_tag": model_tag,
                    "variable": var_name,
                    "term_type": "Linear",
                    "n_basis_functions": 1,
                    "effect_degrees_of_freedom": edof_term,
                    "chi2_statistic": chi2,
                    "f_statistic": fstat,
                    "p_value": pval,
                    "p_value_display": fmt_p(pval),
                }
            )
        elif term_type == "FactorTerm":
            est = np.nan
            se = np.nan
            if cov_mat is not None and feat_idx is not None and feat_idx < len(feature_names):
                try:
                    eff = factor_single_effect(
                        gam=gam,
                        cov_mat=np.asarray(cov_mat, dtype=float),
                        x_ref=x_ref,
                        feat_idx=int(feat_idx),
                        observed_codes=pd.to_numeric(df_ref[feature_names[feat_idx]], errors="coerce").values,
                    )
                    est = float(eff["estimate_beta"])
                    se = float(eff["std_error"])
                except Exception:
                    est = float(c_slice[0]) if len(c_slice) else np.nan
                    se = float(se_slice[0]) if len(se_slice) else np.nan
            else:
                est = float(c_slice[0]) if len(c_slice) else np.nan
                se = float(se_slice[0]) if len(se_slice) else np.nan
            wald_p = wald_p_from_estimate(est, se)
            rows8.append(
                {
                    "ycol": ycol,
                    "cohort": cohort,
                    "model": model_label,
                    "model_tag": model_tag,
                    "variable": var_name,
                    "term_type": "Binary factor",
                    "effect_degrees_of_freedom": 1.0,
                    "estimate_beta": est,
                    "std_error": se,
                    "p_value": wald_p,
                    "p_value_display": fmt_p(wald_p),
                }
            )

        coef_idx += nc

    et7 = pd.DataFrame(rows7)
    et8 = pd.DataFrame(rows8)
    return et6, et7, et8


def write_xlsx(et6: pd.DataFrame, et7: pd.DataFrame, et8: pd.DataFrame, out_xlsx: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    title_font = Font(name="Calibri", size=12, bold=True)
    header_font = Font(name="Calibri", size=10, bold=True)
    body_font = Font(name="Calibri", size=10)
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill = PatternFill(fill_type="solid", fgColor="F2F2F2")
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # eTable 6 (wide)
    ws6 = wb.create_sheet("eTable6_model_fit")
    ws6.merge_cells("A1:D1")
    ws6["A1"] = "eTable 6. Model performance and goodness-of-fit statistics for additive GAMs (no interaction terms)"
    ws6["A1"].font = title_font
    ws6["A1"].alignment = left
    header6 = ["Metric", "Left SctO2 cohort", "Right SctO2 cohort", "SftO2 cohort"]
    for c, v in enumerate(header6, start=1):
        cell = ws6.cell(3, c, v)
        cell.font = header_font
        cell.fill = fill
        cell.border = border
        cell.alignment = left if c == 1 else center

    metric_order = [
        "N Samples",
        "N Features",
        "Effective DOF (model)",
        "Deviance Explained",
        "McFadden R2",
        "AIC",
        "AICc",
        "GCV",
        "UBRE",
        "Log Likelihood",
        "Scale",
        "N Training Rows",
        "RMSE (training sample)",
        "R2 Traditional (training sample)",
        "N Eval Rows (full analytic pool)",
        "RMSE (full analytic pool)",
        "R2 Traditional (full analytic pool)",
    ]
    p6 = et6.pivot_table(index="metric", columns="cohort", values="value", aggfunc="first")
    r = 4
    for metric in metric_order:
        ws6.cell(r, 1, metric).font = body_font
        ws6.cell(r, 1).alignment = left
        ws6.cell(r, 1).border = border
        for j, cohort in enumerate(["Left SctO2 cohort", "Right SctO2 cohort", "SftO2 cohort"], start=2):
            val = p6.loc[metric, cohort] if metric in p6.index and cohort in p6.columns else np.nan
            txt = ""
            if np.isfinite(safe_float(val)):
                txt = f"{float(val):,.6f}" if abs(float(val)) < 1000 else f"{float(val):,.2f}"
            ws6.cell(r, j, txt).font = body_font
            ws6.cell(r, j).alignment = center
            ws6.cell(r, j).border = border
        r += 1
    ws6.column_dimensions["A"].width = 38
    ws6.column_dimensions["B"].width = 24
    ws6.column_dimensions["C"].width = 24
    ws6.column_dimensions["D"].width = 24

    # eTable 7
    ws7 = wb.create_sheet("eTable7_continuous")
    ws7.merge_cells("A1:G1")
    ws7["A1"] = "eTable 7. Smooth term estimates for generalized additive models"
    ws7["A1"].font = title_font
    ws7["A1"].alignment = left
    header7 = ["Model", "Variable", "Term Type", "N basis Functions", "Effect Degrees of Freedom", "F Statistic", "P Value"]
    for c, v in enumerate(header7, start=1):
        cell = ws7.cell(3, c, v)
        cell.font = header_font
        cell.fill = fill
        cell.border = border
        cell.alignment = center if c >= 4 else left
    r = 4
    et7s = et7.copy()
    et7s["model_rank"] = et7s["ycol"].map({k: i for i, k in enumerate(YCOL_ORDER)})
    et7s = et7s.sort_values(["model_rank", "variable"]).drop(columns=["model_rank"])
    for _, row in et7s.iterrows():
        vals = [
            row["model"],
            row["variable"],
            row["term_type"],
            int(row["n_basis_functions"]) if np.isfinite(safe_float(row["n_basis_functions"])) else "",
            f"{float(row['effect_degrees_of_freedom']):.2f}" if np.isfinite(safe_float(row["effect_degrees_of_freedom"])) else "",
            f"{float(row['f_statistic']):.2f}" if np.isfinite(safe_float(row["f_statistic"])) else "",
            row["p_value_display"],
        ]
        for c, v in enumerate(vals, start=1):
            cell = ws7.cell(r, c, v)
            cell.font = body_font
            cell.border = border
            cell.alignment = center if c >= 4 else left
        r += 1
    for col, width in {"A": 24, "B": 18, "C": 18, "D": 18, "E": 24, "F": 14, "G": 10}.items():
        ws7.column_dimensions[col].width = width

    # eTable 8
    ws8 = wb.create_sheet("eTable8_categorical")
    ws8.merge_cells("A1:G1")
    ws8["A1"] = "eTable 8. Binary categorical covariate estimates for generalized additive models"
    ws8["A1"].font = title_font
    ws8["A1"].alignment = left
    header8 = ["Model", "Variable", "Term Type", "Effect Degrees of Freedom", "Estimate(β)", "Std Error", "P Value"]
    for c, v in enumerate(header8, start=1):
        cell = ws8.cell(3, c, v)
        cell.font = header_font
        cell.fill = fill
        cell.border = border
        cell.alignment = center if c >= 4 else left
    r = 4
    et8s = et8.copy()
    et8s["model_rank"] = et8s["ycol"].map({k: i for i, k in enumerate(YCOL_ORDER)})
    et8s = et8s.sort_values(["model_rank", "variable"]).drop(columns=["model_rank"])
    for _, row in et8s.iterrows():
        vals = [
            row["model"],
            row["variable"],
            row["term_type"],
            f"{float(row['effect_degrees_of_freedom']):.0f}" if np.isfinite(safe_float(row["effect_degrees_of_freedom"])) else "",
            f"{float(row['estimate_beta']):.4f}" if np.isfinite(safe_float(row["estimate_beta"])) else "",
            f"{float(row['std_error']):.4f}" if np.isfinite(safe_float(row["std_error"])) else "",
            row["p_value_display"],
        ]
        for c, v in enumerate(vals, start=1):
            cell = ws8.cell(r, c, v)
            cell.font = body_font
            cell.border = border
            cell.alignment = center if c >= 4 else left
        r += 1
    for col, width in {"A": 24, "B": 22, "C": 16, "D": 24, "E": 14, "F": 14, "G": 10}.items():
        ws8.column_dimensions[col].width = width

    wb.save(out_xlsx)


def main():
    if not MODEL_SCRIPT.exists():
        raise FileNotFoundError(f"Model script not found: {MODEL_SCRIPT}")

    m = load_module(MODEL_SCRIPT)
    outcomes_env = [x.strip() for x in os.getenv("ETABLE68_OUTCOMES", "").split(",") if x.strip()]
    ycols = [y for y in outcomes_env if y in YCOL_ORDER] if outcomes_env else list(YCOL_ORDER)
    if len(ycols) < len(YCOL_ORDER):
        print(f"[warn] ETABLE68_OUTCOMES override active, only running cohorts: {ycols}")
    sample_n = int(os.getenv("ETABLE68_SUBSAMPLE_N", str(int(m.SUBSAMPLE_SIZE[0]) if len(m.SUBSAMPLE_SIZE) else 100000)))
    auto_tune = os.getenv("ETABLE68_AUTO_TUNE", "1" if bool(m.AUTO_TUNE_SMOOTH) else "0") == "1"
    sample_method = os.getenv("ETABLE68_SAMPLE_METHOD", str(m.REF_SAMPLE_METHOD))
    replace = os.getenv("ETABLE68_REPLACE", "1" if bool(m.REF_SAMPLE_REPLACE) else "0") == "1"

    et6_all: List[pd.DataFrame] = []
    et7_all: List[pd.DataFrame] = []
    et8_all: List[pd.DataFrame] = []

    print(
        f"[cfg] sample_n={sample_n}, auto_tune={int(auto_tune)}, sample_method={sample_method}, "
        f"replace={int(replace)}, outcomes={ycols}"
    )
    for ycol in ycols:
        print(f"[fit] {ycol}")
        pool = get_cached_pool(m, ycol=ycol, sec=1, subgroup_tag="All", subgroup_query="")
        spec = choose_model_spec(m, pool=pool)
        model_tag = str(spec.get("model_tag", "model"))
        primary_vars = list(spec.get("primary_vars", []))

        d_model, pvars, cont_cov, cat_cov, id_col = m.prepare_model_df(
            pool,
            ycol=ycol,
            extra_exclude_covars=[],
            primary_vars_override=primary_vars,
        )
        if len(d_model) < 500:
            raise RuntimeError(f"Insufficient rows for {ycol}: {len(d_model)}")

        df_ref = m.sample_rows(
            d_model,
            id_col=id_col,
            target_n=int(sample_n),
            replace=replace,
            random_seed=int(m.SEED + 17),
            method=sample_method,
            ensure_exact_n=bool(m.SAMPLE_ENSURE_EXACT_N),
        )
        if len(df_ref) < 500:
            raise RuntimeError(f"Reference sample too small for {ycol}: {len(df_ref)}")

        if auto_tune:
            tuned = m.tune_smoothing(df_ref, ycol=ycol, primary_vars=pvars, cont_cov=cont_cov, cat_cov=cat_cov)
            gam = tuned["gam"]
            n_splines_used = tuned.get("n_splines_main", [])
            lam_used = tuned.get("lam", None)
        else:
            gam = m.fit_gam(
                df_model=df_ref,
                ycol=ycol,
                primary_vars=pvars,
                cont_cov=cont_cov,
                cat_cov=cat_cov,
                n_splines_main=int(m.N_SPLINES_MAIN),
                lam_fixed=float(m.LAM_FIXED),
            )
            n_splines_used = [int(m.N_SPLINES_MAIN)] * len(pvars)
            lam_used = float(m.LAM_FIXED)

        feature_names = m.build_feature_names(pvars, cont_cov, cat_cov)
        cohort = COHORT_MAP.get(ycol, ycol)
        model_label = MODEL_LABEL_MAP.get(ycol, ycol)
        et6, et7, et8 = collect_model_tables(
            gam=gam,
            ycol=ycol,
            cohort=cohort,
            model_label=model_label,
            feature_names=feature_names,
            primary_vars=pvars,
            cont_cov=cont_cov,
            cat_cov=cat_cov,
            df_ref=df_ref,
            df_eval=d_model,
            model_tag=model_tag,
            n_splines_used=n_splines_used,
            lam_used=lam_used,
        )
        et6_all.append(et6)
        et7_all.append(et7)
        et8_all.append(et8)
        print(f"[done] {ycol} rows: et6={len(et6)} et7={len(et7)} et8={len(et8)}")

    d6 = pd.concat(et6_all, ignore_index=True) if et6_all else pd.DataFrame()
    d7 = pd.concat(et7_all, ignore_index=True) if et7_all else pd.DataFrame()
    d8 = pd.concat(et8_all, ignore_index=True) if et8_all else pd.DataFrame()

    d6.to_csv(OUT_ET6, index=False)
    d7.to_csv(OUT_ET7, index=False)
    d8.to_csv(OUT_ET8, index=False)
    write_xlsx(d6, d7, d8, OUT_XLSX)

    print("[write]")
    print(f" - {OUT_ET6}")
    print(f" - {OUT_ET7}")
    print(f" - {OUT_ET8}")
    print(f" - {OUT_XLSX}")


if __name__ == "__main__":
    main()
