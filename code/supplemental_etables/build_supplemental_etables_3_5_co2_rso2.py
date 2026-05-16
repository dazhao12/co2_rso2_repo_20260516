#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Build supplemental eTable 3/4/5 data for CO2-rSO2 project.

Outputs (under GAM_Co2_Scto2_4_19_2026):
  - supplemental_etable3_artifact_co2_rso2.csv
  - supplemental_etable4_missingness_imputation_other_intraop.csv
  - supplemental_etable5_patient_level_co2_rso2.csv
  - Supplemental_eTables3_5_CO2_rSO2.xlsx
"""

import json
import os
from pathlib import Path
from typing import Dict, List, Tuple

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


PROJECT_ROOT = Path("/N/project/waveform_mortality/ZhaoZhang/contour_zhao_all_9_15_2025")
WORK_DIR = PROJECT_ROOT / "GAM_Co2_Scto2_4_19_2026"
RAW_CACHE_PKL_ENV = os.getenv("ETABLE35_RAW_CACHE_PKL", "").strip()
SCRIPT_PY = WORK_DIR / "contour_4_19_2026_intraop10_totalonly_noetco2_subgroup.py"

OUT_ET3 = WORK_DIR / "supplemental_etable3_artifact_co2_rso2.csv"
OUT_ET4 = WORK_DIR / "supplemental_etable4_missingness_imputation_other_intraop.csv"
OUT_ET5 = WORK_DIR / "supplemental_etable5_patient_level_co2_rso2.csv"
OUT_XLSX = WORK_DIR / "Supplemental_eTables3_5_CO2_rSO2.xlsx"

YCOLS = ["rSO2_Ch1", "rSO2_Ch2", "rSO2_Ch3"]
ET4_COVARS = ["TEMP", "FiO2_new", "MAP", "SV", "HR", "CI"]
YLABEL = {
    "rSO2_Ch1": "Left SctO2 cohort",
    "rSO2_Ch2": "Right SctO2 cohort",
    "rSO2_Ch3": "Frontal SctO2 cohort",
}
COHORTS = ["Left SctO2 cohort", "Right SctO2 cohort", "Frontal SctO2 cohort"]

# eTable3 artifact display threshold aligned to primary analytic bounds:
ET_ART_LO = 20.0
ET_ART_HI = 50.0
Y_ART_LO = 20.0
Y_ART_HI = 95.0

# Primary analytic strict threshold in current 4_19 code:
ET_STRICT_LO = 20.0
ET_STRICT_HI = 50.0
Y_STRICT_LO = 20.0
Y_STRICT_HI = 95.0


def med_iqr_str(x: np.ndarray, digits: int = 1) -> str:
    x = x[np.isfinite(x)]
    if x.size == 0:
        return "NA"
    q1, med, q3 = np.percentile(x, [25, 50, 75])
    return f"{med:.{digits}f} ({q1:.{digits}f} – {q3:.{digits}f})"


def mean_sd_str(x: np.ndarray, digits: int = 1) -> str:
    x = x[np.isfinite(x)]
    if x.size == 0:
        return "NA"
    sd = float(np.std(x, ddof=1)) if x.size > 1 else 0.0
    return f"{float(np.mean(x)):.{digits}f} ({sd:.{digits}f})"


def rng_str(x: np.ndarray, digits: int = 1) -> str:
    x = x[np.isfinite(x)]
    if x.size == 0:
        return "NA"
    return f"{np.min(x):.{digits}f} – {np.max(x):.{digits}f}"


def n_pct_str(n: int, den: int) -> str:
    pct = (100.0 * n / den) if den else 0.0
    return f"{n:,} ({pct:.2f}%)"


def _patient_total_missing_minutes(miss: pd.Series, ids: pd.Series, sec_per_point: float = 1.0) -> np.ndarray:
    if len(miss) == 0:
        return np.array([], dtype=float)
    miss_n = miss.astype(bool).groupby(ids).sum().astype(float)
    return (miss_n * float(sec_per_point) / 60.0).to_numpy(dtype=float)


def _gap_durations_seconds(t_vec: np.ndarray, missing_vec: np.ndarray) -> np.ndarray:
    t = pd.to_numeric(pd.Series(t_vec), errors="coerce").to_numpy(dtype=float)
    m = np.asarray(missing_vec, dtype=bool)
    if t.size == 0 or t.size != m.size:
        return np.array([], dtype=float)
    ok = np.isfinite(t)
    t = t[ok]
    m = m[ok]
    if t.size == 0:
        return np.array([], dtype=float)

    dt = np.diff(t)
    dt = dt[np.isfinite(dt) & (dt > 0)]
    dt_med = float(np.median(dt)) if dt.size else 0.0

    r = pd.Series(m).astype(bool).to_numpy()
    if not r.any():
        return np.array([], dtype=float)

    # run-length boundaries for True runs
    change = np.diff(np.r_[False, r, False].astype(int))
    starts = np.where(change == 1)[0]
    ends = np.where(change == -1)[0] - 1
    if starts.size == 0 or ends.size == 0:
        return np.array([], dtype=float)

    dur = (t[ends] - t[starts]) + dt_med
    dur = dur[np.isfinite(dur) & (dur >= 0)]
    return dur.astype(float)


def _gap_missing_minutes_all_segments(miss: pd.Series, times: pd.Series, ids: pd.Series) -> np.ndarray:
    if len(miss) == 0:
        return np.array([], dtype=float)
    id_arr = ids.astype(str).str.strip().to_numpy(dtype=object)
    t_arr = pd.to_numeric(times, errors="coerce").to_numpy(dtype=float)
    m_arr = miss.astype(bool).to_numpy(dtype=bool)

    valid_id = np.fromiter(
        ((x != "") and (str(x).lower() != "nan") for x in id_arr),
        dtype=bool,
        count=id_arr.shape[0],
    )
    valid = valid_id & np.isfinite(t_arr)
    if not np.any(valid):
        return np.array([], dtype=float)

    id_arr = id_arr[valid]
    t_arr = t_arr[valid]
    m_arr = m_arr[valid]

    order = np.lexsort((t_arr, id_arr))
    id_arr = id_arr[order]
    t_arr = t_arr[order]
    m_arr = m_arr[order]

    # contiguous segments per patient id after sorting
    starts = np.r_[0, np.where(id_arr[1:] != id_arr[:-1])[0] + 1]
    ends = np.r_[starts[1:], id_arr.shape[0]]

    chunks: List[np.ndarray] = []
    for s, e in zip(starts, ends):
        dur_sec = _gap_durations_seconds(t_arr[s:e], m_arr[s:e])
        if dur_sec.size > 0:
            chunks.append(dur_sec / 60.0)

    if len(chunks) == 0:
        return np.array([], dtype=float)
    return np.concatenate(chunks).astype(float)


def pick_id_col(df: pd.DataFrame) -> str:
    for c in ["stay_id", "patient_ID", "patient_id"]:
        if c in df.columns:
            return c
    raise ValueError("No patient id column found in dataframe.")


def load_py_module(script_path: Path):
    import importlib.util

    spec = importlib.util.spec_from_file_location("m4_19", script_path)
    m = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(m)  # type: ignore[attr-defined]
    return m


def load_analysis_base_df(m) -> pd.DataFrame:
    # eTable4 needs dynamic covariates; keep these columns in the base load.
    need_cols = ["stay_id", "patient_ID", "patient_id", "obstime", "ET_CO2"] + YCOLS + ET4_COVARS
    df = m.read_csv_folder_selected_cached(m.CSV_DIR, m.CSV_GLOB, need_cols=need_cols)
    for c in ["ET_CO2"] + YCOLS + ET4_COVARS:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce")
    return df


def build_etable3(df: pd.DataFrame) -> pd.DataFrame:
    out = []
    et = pd.to_numeric(df["ET_CO2"], errors="coerce")
    id_col = pick_id_col(df)

    for y in YCOLS:
        yy = pd.to_numeric(df[y], errors="coerce")
        nonmiss = et.notna() & yy.notna()
        den_points = int(nonmiss.sum())
        et_nm = et[nonmiss].to_numpy(dtype=float)
        yy_nm = yy[nonmiss].to_numpy(dtype=float)
        id_s = df.loc[nonmiss, id_col].astype(str).str.strip()
        valid_id_s = (id_s != "") & (id_s.str.lower() != "nan")
        id_nm = id_s.to_numpy()
        valid_id = valid_id_s.to_numpy()
        den_patients = int(id_s[valid_id_s].nunique())

        criteria = [
            ("EtCO2 ≤20", et_nm <= ET_ART_LO),
            ("EtCO2 ≥50", et_nm >= ET_ART_HI),
            ("Tissue oxygen ≤20", yy_nm <= Y_ART_LO),
            ("Tissue oxygen ≥95", yy_nm >= Y_ART_HI),
        ]

        for crit, m in criteria:
            n = int(np.sum(m))
            et_rm = et_nm[m]
            y_rm = yy_nm[m]
            pats_rm = int(pd.Series(id_nm[m & valid_id]).nunique()) if den_patients else 0
            out.append(
                {
                    "cohort": YLABEL[y],
                    "ycol": y,
                    "available_data_points_n": den_points,
                    "available_patients_n": den_patients,
                    "outlier_criterion": crit,
                    "data_points_removed_n_pct": n_pct_str(n, den_points),
                    "removed_n": n,
                    "removed_pct": (100.0 * n / den_points) if den_points else 0.0,
                    "patients_with_removed_points_n_pct": n_pct_str(pats_rm, den_patients),
                    "patients_with_removed_points_n": pats_rm,
                    "patients_with_removed_points_pct": (100.0 * pats_rm / den_patients) if den_patients else 0.0,
                    "removed_EtCO2_median_iqr_mmHg": med_iqr_str(et_rm, digits=1),
                    "removed_EtCO2_range_mmHg": rng_str(et_rm, digits=1),
                    "corresponding_tissue_oxygen_median_iqr_pct": med_iqr_str(y_rm, digits=1),
                    "corresponding_tissue_oxygen_range_pct": rng_str(y_rm, digits=1),
                }
            )
    return pd.DataFrame(out)


def build_etable4(df: pd.DataFrame, m) -> pd.DataFrame:
    """
    Read flow/clip/fill summaries from processed pool cache for current 4_19 config.
    """
    data_source_sig = m.build_data_source_signature()
    out = []
    missing_cache_dirs: List[str] = []
    id_col = pick_id_col(df)
    vars_interest = ET4_COVARS
    vars_present = [v for v in vars_interest if v in df.columns]

    for y in YCOLS:
        key = m.build_pool_cache_key(
            data_source_sig=data_source_sig,
            sec=1,
            subgroup_tag="All",
            subgroup_query="",
            ycol=y,
        )
        cache_dir = Path(m.POOL_CACHE_ROOT) / f"intra5_pool_{key}"
        if not cache_dir.exists():
            missing_cache_dirs.append(str(cache_dir))
            continue

        flow_fp = cache_dir / "flow_rows.json"
        clip_fp = cache_dir / "clip_rows.json"
        fill_fp = cache_dir / "fill_rows.json"
        missing_files = [str(p) for p in [flow_fp, clip_fp, fill_fp] if not p.exists()]
        if missing_files:
            raise FileNotFoundError(
                "Missing required eTable4 cache summary files:\n"
                + "\n".join(missing_files)
            )

        flow = json.loads(flow_fp.read_text(encoding="utf-8"))
        clip = json.loads(clip_fp.read_text(encoding="utf-8"))
        fill = json.loads(fill_fp.read_text(encoding="utf-8"))

        flow_map = {r["stage"]: r for r in flow}
        stage_key = "final_usable_points_strict_etco2_rso2"
        if stage_key not in flow_map:
            raise KeyError(f"Stage '{stage_key}' not found in {flow_fp}")
        den = int(flow_map[stage_key]["n_rows"])

        clip_map = {r["column"]: r for r in clip}
        fill_map = {r["column"]: r for r in fill}

        et = pd.to_numeric(df["ET_CO2"], errors="coerce")
        yy = pd.to_numeric(df[y], errors="coerce")
        strict = (
            et.notna()
            & yy.notna()
            & (et > ET_STRICT_LO)
            & (et < ET_STRICT_HI)
            & (yy > Y_STRICT_LO)
            & (yy < Y_STRICT_HI)
        )
        d = df.loc[strict, [id_col, "obstime"] + vars_present].copy()
        d["obstime"] = pd.to_numeric(d["obstime"], errors="coerce")

        id_clean = d[id_col].astype(str).str.strip()
        valid_id = id_clean.notna() & (id_clean != "") & (id_clean.str.lower() != "nan")
        d_pat = d.loc[valid_id, [id_col] + vars_present].copy()
        d_pat[id_col] = d_pat[id_col].astype(str).str.strip()

        # Old-R-like monitor-axis basis for gap-level missing episode duration:
        # same patient cohort, but across all available monitor rows in base df.
        cohort_ids = set(d_pat[id_col].astype(str).tolist())
        id_all = df[id_col].astype(str).str.strip()
        valid_all = (id_all != "") & id_all.notna() & (id_all.str.lower() != "nan")
        in_cohort = valid_all & id_all.isin(cohort_ids)
        d_monitor = df.loc[in_cohort, [id_col, "obstime"] + vars_present].copy()
        d_monitor[id_col] = id_all.loc[in_cohort].to_numpy()
        d_monitor["obstime"] = pd.to_numeric(d_monitor["obstime"], errors="coerce")

        for v in vars_interest:
            c = clip_map.get(v, {})
            f = fill_map.get(v, {})

            s_all = pd.to_numeric(d[v], errors="coerce") if v in d.columns else pd.Series(np.nan, index=d.index)
            lo_hi = m.PHYSIO_CLIP_RANGES.get(v, None)
            if lo_hi is not None:
                lo, hi = lo_hi
                s_all = s_all.where(s_all.between(lo, hi, inclusive="both"))
            nonmiss_vals = s_all.dropna().to_numpy(dtype=float)
            missing_after_clip = int(s_all.isna().sum())

            ff = int(f.get("filled_by_forward_fill", 0))
            sm = int(f.get("filled_by_subject_median", 0))
            gm = int(f.get("filled_by_global_median", 0))
            median_imp_n = sm + gm

            if v in d_pat.columns and len(d_pat):
                s_pat = pd.to_numeric(d_pat[v], errors="coerce")
                if lo_hi is not None:
                    lo, hi = lo_hi
                    s_pat = s_pat.where(s_pat.between(lo, hi, inclusive="both"))
                miss = s_pat.isna()
                dur_min_all = _patient_total_missing_minutes(miss=miss, ids=d_pat[id_col], sec_per_point=1.0)
                miss_dur_all_patients_median_iqr = med_iqr_str(dur_min_all, digits=0)
                dur_min_missing_only = dur_min_all[dur_min_all > 0]
                miss_dur_missing_patients_median_iqr = med_iqr_str(dur_min_missing_only, digits=0)
            else:
                miss_dur_all_patients_median_iqr = "NA"
                miss_dur_missing_patients_median_iqr = "NA"

            # Main old-method metric: episode-level missing duration on monitor axis.
            if v in d_monitor.columns and len(d_monitor):
                s_mon = pd.to_numeric(d_monitor[v], errors="coerce")
                if lo_hi is not None:
                    lo, hi = lo_hi
                    s_mon = s_mon.where(s_mon.between(lo, hi, inclusive="both"))
                miss_mon = s_mon.isna()
                gap_min = _gap_missing_minutes_all_segments(
                    miss=miss_mon, times=d_monitor["obstime"], ids=d_monitor[id_col]
                )
                miss_episode_dur_median_iqr = med_iqr_str(gap_min, digits=0)
            else:
                miss_episode_dur_median_iqr = "NA"

            out.append(
                {
                    "cohort": YLABEL[y],
                    "ycol": y,
                    "covariate": v,
                    "available_data_points_n": den,
                    "mean_sd": mean_sd_str(nonmiss_vals, digits=1),
                    "median_iqr": med_iqr_str(nonmiss_vals, digits=1),
                    "missing_after_clip_n": missing_after_clip,
                    "missing_after_clip_pct": (100.0 * missing_after_clip / den) if den else 0.0,
                    "forward_imputation_n": ff,
                    "forward_imputation_pct": (100.0 * ff / den) if den else 0.0,
                    "median_imputation_n": median_imp_n,
                    "median_imputation_pct": (100.0 * median_imp_n / den) if den else 0.0,
                    "median_missing_episode_duration_iqr_min": miss_episode_dur_median_iqr,
                    "median_missing_duration_all_patients_iqr_min": miss_dur_all_patients_median_iqr,
                    "median_missing_duration_missing_patients_iqr_min": miss_dur_missing_patients_median_iqr,
                    "newly_clipped_to_missing_n": int(c.get("newly_clipped_to_missing", 0)),
                }
            )

    if missing_cache_dirs:
        msg = (
            "Missing processed pool cache directories for eTable4 (one per cohort). "
            "Run contour_4_19_2026_intraop10_totalonly_noetco2_subgroup.py first, then rerun this script.\n"
            + "\n".join(missing_cache_dirs)
        )
        raise FileNotFoundError(msg)

    d = pd.DataFrame(out)
    if len(d):
        d["missing_after_clip_pct"] = d["missing_after_clip_pct"].round(2)
        d["forward_imputation_pct"] = d["forward_imputation_pct"].round(2)
        d["median_imputation_pct"] = d["median_imputation_pct"].round(2)
    return d


def build_etable5(df: pd.DataFrame) -> pd.DataFrame:
    out = []
    id_col = pick_id_col(df)
    et = pd.to_numeric(df["ET_CO2"], errors="coerce")

    for y in YCOLS:
        yy = pd.to_numeric(df[y], errors="coerce")
        m = (
            et.notna()
            & yy.notna()
            & (et > ET_STRICT_LO)
            & (et < ET_STRICT_HI)
            & (yy > Y_STRICT_LO)
            & (yy < Y_STRICT_HI)
        )

        d = df.loc[m, [id_col]].copy()
        d["ET_CO2"] = et[m].to_numpy(dtype=float)
        d[y] = yy[m].to_numpy(dtype=float)
        d[id_col] = d[id_col].astype(str).str.strip()
        d = d[(d[id_col] != "") & d[id_col].notna()]

        if len(d) == 0:
            continue

        g = d.groupby(id_col, as_index=False).agg(
            n_points=("ET_CO2", "size"),
            et_mean=("ET_CO2", "mean"),
            et_sd=("ET_CO2", "std"),
            y_mean=(y, "mean"),
            y_sd=(y, "std"),
        )
        for c in ["et_sd", "y_sd"]:
            g[c] = g[c].fillna(0.0)

        q = g["n_points"].quantile([0.25, 0.5, 0.75]).to_dict()

        out.append(
            {
                "cohort": YLABEL[y],
                "ycol": y,
                "patients_with_recordings_n": int(g.shape[0]),
                "median_data_points_per_patient_iqr_n": f"{q[0.5]:.0f} ({q[0.25]:.0f} – {q[0.75]:.0f})",
                "mean_of_patient_mean_EtCO2_sd_mmHg": f"{g['et_mean'].mean():.1f} ({g['et_mean'].std(ddof=1):.1f})",
                "mean_of_patient_sd_EtCO2_sd_mmHg": f"{g['et_sd'].mean():.1f} ({g['et_sd'].std(ddof=1):.1f})",
                "mean_of_patient_mean_tissue_oxygen_sd_pct": f"{g['y_mean'].mean():.1f} ({g['y_mean'].std(ddof=1):.1f})",
                "mean_of_patient_sd_tissue_oxygen_sd_pct": f"{g['y_sd'].mean():.1f} ({g['y_sd'].std(ddof=1):.1f})",
                "n_points_total": int(len(d)),
            }
        )
    return pd.DataFrame(out)


def fmt_int(x) -> str:
    if x is None or x == "" or pd.isna(x):
        return ""
    return f"{int(round(float(x))):,}"


def fmt_pct(x, digits: int = 2) -> str:
    if x is None or x == "" or pd.isna(x):
        return ""
    return f"{float(x):.{digits}f}"


def fmt_num(x, digits: int = 2) -> str:
    if x is None or x == "" or pd.isna(x):
        return ""
    return f"{float(x):.{digits}f}"


def _cohort_vals(sub: pd.DataFrame, col: str, fmt=None) -> List[str]:
    vals = []
    for cohort in COHORTS:
        v = sub.loc[cohort, col] if cohort in sub.index and col in sub.columns else ""
        if fmt is None:
            vals.append("" if pd.isna(v) else str(v))
        else:
            vals.append(fmt(v))
    return vals


def format_etable3_rows(et3: pd.DataFrame) -> List[Tuple[List[str], str]]:
    rows: List[Tuple[List[str], str]] = []
    base = et3.drop_duplicates("cohort").set_index("cohort")
    rows.append(
        (
            ["Available data points, n*"] + [fmt_int(base.loc[c, "available_data_points_n"]) for c in COHORTS],
            "data",
        )
    )
    rows.append(
        (
            ["Available patients, n*"] + [fmt_int(base.loc[c, "available_patients_n"]) for c in COHORTS],
            "data",
        )
    )

    criteria_order = [
        "EtCO2 ≤20",
        "EtCO2 ≥50",
        "Tissue oxygen ≤20",
        "Tissue oxygen ≥95",
    ]
    for crit in criteria_order:
        sub = et3.loc[et3["outlier_criterion"] == crit].set_index("cohort")
        rows.append(([crit, crit, crit, crit], "section"))
        rows.append((["Data points removed, n (%)"] + _cohort_vals(sub, "data_points_removed_n_pct"), "data"))
        rows.append((["Patients with removed points, n (%)"] + _cohort_vals(sub, "patients_with_removed_points_n_pct"), "data"))
        rows.append((["Median EtCO2 (IQR), mmHg"] + _cohort_vals(sub, "removed_EtCO2_median_iqr_mmHg"), "data"))
        rows.append((["EtCO2 range, mmHg"] + _cohort_vals(sub, "removed_EtCO2_range_mmHg"), "data"))
        rows.append(
            (
                ["Median corresponding tissue oxygen (IQR), %"]
                + _cohort_vals(sub, "corresponding_tissue_oxygen_median_iqr_pct"),
                "data",
            )
        )
        rows.append(
            (
                ["Corresponding tissue oxygen range, %"] + _cohort_vals(sub, "corresponding_tissue_oxygen_range_pct"),
                "data",
            )
        )
    return rows


def format_etable4_rows(et4: pd.DataFrame) -> List[Tuple[List[str], str]]:
    rows: List[Tuple[List[str], str]] = []
    base = et4.drop_duplicates("cohort").set_index("cohort")
    rows.append(
        (
            ["Available data points, n*"] + [fmt_int(base.loc[c, "available_data_points_n"]) for c in COHORTS],
            "data",
        )
    )

    cov_order = [
        ("TEMP", "Temperature", "temperature", "°C"),
        ("FiO2_new", "FiO2", "FiO2", "%"),
        ("MAP", "MAP", "MAP", "mmHg"),
        ("SV", "SV", "SV", "mL"),
        ("HR", "HR", "HR", "beats/min"),
        ("CI", "CI", "CI", "L/min/m²"),
    ]
    for cov, label, name_text, unit in cov_order:
        sub = et4.loc[et4["covariate"] == cov].set_index("cohort")
        rows.append(([label, label, label, label], "section"))
        rows.append(([f"Mean {name_text} (SD), {unit}"] + _cohort_vals(sub, "mean_sd"), "data"))
        rows.append(([f"Median {name_text} (IQR), {unit}"] + _cohort_vals(sub, "median_iqr"), "data"))
        rows.append(([f"Data points with {name_text} missing, n"] + _cohort_vals(sub, "missing_after_clip_n", fmt_int), "data"))
        rows.append((["Missing rate, %"] + _cohort_vals(sub, "missing_after_clip_pct", lambda x: fmt_pct(x, 1)), "data"))
        rows.append((["Forward imputation, %"] + _cohort_vals(sub, "forward_imputation_pct", lambda x: fmt_pct(x, 1)), "data"))
        rows.append((["Median imputation, %"] + _cohort_vals(sub, "median_imputation_pct", lambda x: fmt_pct(x, 1)), "data"))
        rows.append(
            (
                ["Median duration of missingness by episode (IQR), min†"]
                + _cohort_vals(sub, "median_missing_episode_duration_iqr_min"),
                "data",
            )
        )
        rows.append(
            (
                ["Median total missing duration across all patients (IQR), min‡"]
                + _cohort_vals(sub, "median_missing_duration_all_patients_iqr_min"),
                "data",
            )
        )
        rows.append(
            (
                ["Median total missing duration among patients with any missingness (IQR), min§"]
                + _cohort_vals(sub, "median_missing_duration_missing_patients_iqr_min"),
                "data",
            )
        )
    return rows


def format_etable5_rows(et5: pd.DataFrame) -> List[Tuple[List[str], str]]:
    rows: List[Tuple[List[str], str]] = []
    sub = et5.set_index("cohort")
    rows.append((["Patients with recordings, n"] + _cohort_vals(sub, "patients_with_recordings_n", fmt_int), "data"))
    rows.append((["Total included data points, n*"] + _cohort_vals(sub, "n_points_total", fmt_int), "data"))
    rows.append(
        (
            ["Median data points per patient (IQR), n"]
            + _cohort_vals(sub, "median_data_points_per_patient_iqr_n"),
            "data",
        )
    )
    rows.append(
        (
            ["Mean of each patient's mean EtCO2 (SD), mmHg†"]
            + _cohort_vals(sub, "mean_of_patient_mean_EtCO2_sd_mmHg"),
            "data",
        )
    )
    rows.append(
        (
            ["Mean of SD of each patient's EtCO2 (SD), mmHg‡"]
            + _cohort_vals(sub, "mean_of_patient_sd_EtCO2_sd_mmHg"),
            "data",
        )
    )
    rows.append(
        (
            ["Mean of each patient's mean tissue oxygenation (SD), %†"]
            + _cohort_vals(sub, "mean_of_patient_mean_tissue_oxygen_sd_pct"),
            "data",
        )
    )
    rows.append(
        (
            ["Mean of SD of each patient's tissue oxygenation (SD), %‡"]
            + _cohort_vals(sub, "mean_of_patient_sd_tissue_oxygen_sd_pct"),
            "data",
        )
    )
    return rows


def write_sheet(
    wb: Workbook,
    sheet_name: str,
    title: str,
    first_col_header: str,
    rows: List[Tuple[List[str], str]],
    abbreviations: str,
    footnotes: List[str],
) -> None:
    ws = wb.create_sheet(title=sheet_name)

    title_font = Font(name="Calibri", size=12, bold=True)
    header_font = Font(name="Calibri", size=10, bold=True)
    body_font = Font(name="Calibri", size=10)
    note_font = Font(name="Calibri", size=10)
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill(fill_type="solid", fgColor="F2F2F2")
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.merge_cells("A1:D1")
    ws["A1"] = title
    ws["A1"].font = title_font
    ws["A1"].alignment = align_left

    header_row = 3
    header = [first_col_header] + COHORTS
    for col, val in enumerate(header, start=1):
        cell = ws.cell(row=header_row, column=col, value=val)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = align_left if col == 1 else align_center

    row_idx = header_row + 1
    for values, kind in rows:
        is_section = kind == "section"
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font = Font(name="Calibri", size=10, bold=is_section)
            cell.border = border
            cell.alignment = align_left if col == 1 else align_center
        row_idx += 1

    ws.row_dimensions[1].height = 30
    for r in range(header_row, row_idx):
        ws.row_dimensions[r].height = 24

    ws.column_dimensions["A"].width = 58
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 26
    ws.column_dimensions["D"].width = 26
    ws.freeze_panes = "B4"

    note_row = row_idx + 1
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=4)
    ws.cell(row=note_row, column=1, value=abbreviations).font = note_font
    ws.cell(row=note_row, column=1).alignment = align_left

    for fn in footnotes:
        note_row += 1
        ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=4)
        c = ws.cell(row=note_row, column=1, value=fn)
        c.font = note_font
        c.alignment = align_left


def write_formatted_workbook(et3: pd.DataFrame, et4: pd.DataFrame, et5: pd.DataFrame, out_xlsx: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    write_sheet(
        wb=wb,
        sheet_name="eTable3_artifact",
        title=(
            "eTable 3. Summary of timestamp-level outlier exclusions in left cerebral, right cerebral, "
            "and frontal cerebral tissue oxygen saturation cohorts"
        ),
        first_col_header="Outlier criterion",
        rows=format_etable3_rows(et3),
        abbreviations=(
            "SctO2, cerebral tissue oxygen saturation; EtCO2, end-tidal carbon dioxide."
        ),
        footnotes=[
            (
                "* Available data points represent synchronized observations with non-missing EtCO2 and "
                "site-specific tissue oxygen saturation."
            ),
            (
                "Thresholds used for artifact reporting: EtCO2 ≤20 or ≥50 mmHg; tissue oxygen saturation ≤20% or ≥95%."
            ),
        ],
    )

    write_sheet(
        wb=wb,
        sheet_name="eTable4_imputation",
        title=(
            "eTable 4. Missingness and imputation of intraoperative time-varying covariates in left cerebral, "
            "right cerebral, and frontal cerebral tissue oxygenation cohorts"
        ),
        first_col_header="Covariate",
        rows=format_etable4_rows(et4),
        abbreviations=(
            "SctO2, cerebral tissue oxygen saturation; EtCO2, end-tidal carbon dioxide; "
            "FiO2, fraction of inspired oxygen."
        ),
        footnotes=[
            (
                "* Available data points represent analytic observations after excluding timestamps with EtCO2 "
                "or site-specific tissue oxygen artifact values."
            ),
            "Percentages are calculated using available data points in each cohort as the denominator.",
            "Median imputation combines subject-level median and global median imputation.",
            "† Values represent missing-episode-level duration summary based on monitor-axis timestamps in the cohort patients.",
            "‡ Values represent the median of patient-level total missing duration, including patients with 0 minutes missing.",
            "§ Values represent the median of patient-level total missing duration among patients with any missing data (>0 minutes).",
        ],
    )

    write_sheet(
        wb=wb,
        sheet_name="eTable5_patient_level",
        title=(
            "eTable 5. Patient-level summary of intraoperative EtCO2 and tissue oxygenation in left cerebral, "
            "right cerebral, and frontal cerebral tissue oxygenation cohorts"
        ),
        first_col_header="Variable",
        rows=format_etable5_rows(et5),
        abbreviations="SctO2, cerebral tissue oxygen saturation.",
        footnotes=[
            "* Included data points met strict thresholds: 20<EtCO2<50 mmHg and 20<tissue oxygen saturation<95%.",
            "† Mean of patient-level means over the intraoperative period.",
            "‡ Mean of patient-level SDs over the intraoperative period.",
        ],
    )

    wb.save(out_xlsx)


def main():
    if not SCRIPT_PY.exists():
        raise FileNotFoundError(f"4_19 script not found: {SCRIPT_PY}")

    m = load_py_module(SCRIPT_PY)
    if RAW_CACHE_PKL_ENV:
        raw_cache_pkl = Path(RAW_CACHE_PKL_ENV).expanduser().resolve()
        if not raw_cache_pkl.exists():
            raise FileNotFoundError(f"ETABLE35_RAW_CACHE_PKL not found: {raw_cache_pkl}")
        print(f"[load] raw cache (override): {raw_cache_pkl}")
        df = pd.read_pickle(raw_cache_pkl)
        for c in ["ET_CO2"] + YCOLS:
            if c in df.columns:
                df[c] = pd.to_numeric(df[c], errors="coerce")
    else:
        print("[load] dynamic source via current contour script config")
        df = load_analysis_base_df(m)

    print("[build] eTable3")
    et3 = build_etable3(df)
    et3.to_csv(OUT_ET3, index=False)

    print("[build] eTable4")
    et4 = build_etable4(df, m)
    et4.to_csv(OUT_ET4, index=False)

    print("[build] eTable5")
    et5 = build_etable5(df)
    et5.to_csv(OUT_ET5, index=False)

    print(f"[write] xlsx: {OUT_XLSX}")
    write_formatted_workbook(et3=et3, et4=et4, et5=et5, out_xlsx=OUT_XLSX)

    print("[done]")
    print(f" - {OUT_ET3}")
    print(f" - {OUT_ET4}")
    print(f" - {OUT_ET5}")
    print(f" - {OUT_XLSX}")


if __name__ == "__main__":
    main()
